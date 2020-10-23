import tkinter as tk
from datetime import datetime, date
import re
import pandas
import tkinter.font as font
from openpyxl import load_workbook
import random


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.timeIN = 0
        self.timeOut = 0
        self.hoursWorked = 0
        self.minsWorked = 0
        self.date = 0
        self.master = master
        self.master.configure(background='black')
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        myFont = font.Font(family='Courier',
                           size=20,
                           weight='bold')

        self.forgotClockIn = tk.Entry(self,
                                      width=16,
                                      bg="#686664",
                                      fg="white",
                                      font=myFont,
                                      justify='center',
                                      borderwidth=5,
                                      relief="ridge")

        self.forgotClockIn.grid(row=1, column=0)
        self.forgotClockIn.insert('end', "00:00")
        self.forgotClockIn.grid_forget()

        self.header = tk.Label(self,
                               font=myFont,
                               text="CLOCK IN\n\nCLOCK OUT",
                               fg='white',
                               bg="#2e3047",
                               height=5,
                               width=16,
                               borderwidth=5,
                               relief="ridge")

        self.header.grid(row=0, column=0)

        self.ClockInImage = tk.PhotoImage(file=r".\Clock.png")
        self.ClockInImage = self.ClockInImage.subsample(2, 2)

        self.timeIn = tk.Button(self,
                                text="TIME IN",
                                font=myFont,
                                image=self.ClockInImage,
                                compound=tk.TOP,
                                command=self.punchIN,
                                fg='black',
                                bg="#2e3047",
                                activebackground="#686664",
                                borderwidth=5,
                                relief="ridge")

        self.timeIn.grid(row=2, column=0)

        self.timeOut = tk.Button(self,
                                 text="TIME OUT",
                                 font=myFont,
                                 image=self.ClockInImage,
                                 fg='black',
                                 bg="#2e3047",
                                 compound=tk.TOP,
                                 command=self.punchOUT,
                                 activebackground="#686664",
                                 borderwidth=5,
                                 relief="ridge")

        self.timeOut.grid(row=3, column=0)

    def rounded(self, val):
        if val < 25:
            return 25
        elif val < 50:
            return 50
        elif val < 75:
            return 75
        else:
            self.hoursWorked += 1
            return 0

    def punchIN(self):
        t = datetime.now()
        self.timeIN = t.strftime("%H:%M:%S")

    def deletePreText(self, idk):
        self.forgotClockIn.delete(0, 'end')

    def punchOUT(self):
        t = datetime.now()
        self.timeOut = t.strftime("%H:%M:%S")
        if self.timeIN == 0:
            root.geometry('280x840')
            root.bind('<Return>', self.fix)
            root.bind('<Button-1>', self.deletePreText)
            self.header['text'] = "You Forgot\nTo Clock In"
            self.forgotClockIn.grid(row=1, column=0)
        else:
            self.cal(self.timeIN, self.timeOut)

    def fix(self, idk):
        valid = True
        adjust = self.forgotClockIn.get()

        if len(adjust) <= 4:
            adjust = '0' + adjust

        for c in adjust:
            if not c.isdigit() and c != ':':
                valid = False
                self.header['text'] = "Not A Valid Time\nUse This Format\n00:00"
                self.header['fg'] = "#" + str(random.randint(100000, 999999))
                self.forgotClockIn.delete(0, 'end')

        if valid:
            self.timeIN = adjust[:2] + ':' + adjust[2:] + ':00'
            self.cal(self.timeIN, self.timeOut)

    def cal(self, timein, timeout):
        tempout = int(re.sub('[^0-9]', '', timeout))
        tempin = int(re.sub('[^0-9]', '', timein))
        tempdif = str(tempout - tempin)
        self.hoursWorked = int(tempdif[:1])
        self.minsWorked = int(tempdif[1:3]) * 100 / 60
        self.minsWorked = self.rounded(self.minsWorked)
        self.header["text"] = "YOU HAVE WORKED\n" + str(self.hoursWorked) + ':' + str(self.minsWorked)
        self.appendToXlsx(self.timeIN, self.timeOut, self.hoursWorked, self.minsWorked)

    def appendToXlsx(idk, In, Out, Hours, Mins, startrow=None, sheet_name='Sheet1', **to_excel_kwargs):
        path = r'C:\Users\Keith Rich\Documents\PythonScripts\TimeCard\TimeInTimeOut.xlsx'
        book = load_workbook(path)
        writer = pandas.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        #df = {"IN": [In], "OUT": [Out], "HOURS:MINS": [str(Hours) + ":" + str(Mins)]}
        hm = str(Hours) + ":" + str(Mins)
        df = [[In, Out, hm]]
        info = pandas.DataFrame(data=df, columns=['IN', 'OUT', 'HOURS:MINS'], index=[str(datetime.date(datetime.now()))])

        info.to_excel(writer, 'Sheet1', startrow=startrow, header= False, **to_excel_kwargs)
        writer.save()
        writer.close()


root = tk.Tk()
root.geometry('280x800')
app = Application(master=root)
app.mainloop()
